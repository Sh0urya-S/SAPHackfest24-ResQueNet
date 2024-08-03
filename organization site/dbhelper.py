import openpyxl

class DBhelper:
    def register(self,ngoName,checkboxs,phn_no,goal,fundingSource,achievements,password):
        filename = 'ngologin.xlsx'
        num_people = (ngoName, checkboxs, phn_no, goal, fundingSource, achievements, password)
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}
        expected_headers = set(['name', 'services', 'phone', 'goal', 'sourceoffunding', 'Achievement', 'password'])
        missing_headers = expected_headers - set(headers.keys())
        if missing_headers:
            print(f"Error: Missing headers: {', '.join(missing_headers)}")
            raise Exception("Missing headers")
        Name = headers.get('name')
        checkbox = headers.get('services')
        phn_nos = headers.get('phone')
        goal = headers.get('goal')
        fundingSource = headers.get('sourceoffunding')
        achievements = headers.get('Achievement')
        password = headers.get('password')
        if any(col is None for col in [Name, checkbox, phn_nos, goal, fundingSource, achievements, password]):
            print(f"Error: Some columns are missing in the header row.")
            raise Exception("Missing columns")

        last_row = sheet.max_row + 1
        sheet.cell(row=last_row, column=Name).value = num_people[0]
        sheet.cell(row=last_row, column=checkbox).value = num_people[1]
        sheet.cell(row=last_row, column=phn_nos).value = num_people[2]
        sheet.cell(row=last_row, column=goal).value = num_people[3]
        sheet.cell(row=last_row, column=fundingSource).value = num_people[4]
        sheet.cell(row=last_row, column=achievements).value = num_people[5]
        sheet.cell(row=last_row, column=password).value = num_people[6]
        workbook.save(filename)
        print("Data appended successfully!")
        return 1

    def search(self,phone,password):
        workbook = openpyxl.load_workbook("ngologin.xlsx")
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2):
            phone_cell = row[2]
            password_cell = row[6]
            if phone_cell.value == phone and password_cell.value == password:
                return 1

    def require(self,phone,password):
        workbook = openpyxl.load_workbook("ngologin.xlsx")
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2):
            phone_cell = row[2]
            password_cell = row[6]
            if phone_cell.value == phone and password_cell.value == password:
                other_data = [cell.value for cell in row[0:6] + row[7:]]
                return other_data[0],other_data[1]

    def select(self,require):
        workbook = openpyxl.load_workbook("user.xlsx")
        worksheet = workbook.active
        result=[]
        for row in worksheet.iter_rows(min_row=2):
            if require=="food&nutrition":
                check = row[4]
            if require=="water":
                check = row[5]
            if require=="medical":
                check = row[6]
            if require=="sanitary":
                check = row[7]
            if check.value == 1:
                other_data = [cell.value for cell in row[0:]]
                result.append(other_data)
        print(result)
        return result

    def emergency(self):
        workbook = openpyxl.load_workbook("emergency.xlsx")
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=2):
            row_data = [cell.value for cell in row]
            data.append(row_data)
        return data


    def essential(self,username, noofpeople,latitude,longitude,foodnutrition,water,medical,sanitary):
        print("hello",username, noofpeople,latitude,longitude,foodnutrition,water,medical,sanitary,"Pending")
        filename = 'user.xlsx'
        num_people = (username, noofpeople,latitude,longitude,foodnutrition,water,medical,sanitary,"Pending")
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            name = [col for col in sheet[1] if col.value == 'username'][0].column
            people = [col for col in sheet[1] if col.value == 'noofpeople'][0].column
            lat = [col for col in sheet[1] if col.value == 'Latitude'][0].column
            lng = [col for col in sheet[1] if col.value == 'Longitude'][0].column
            food = [col for col in sheet[1] if col.value == 'Food&Nutrition'][0].column
            water = [col for col in sheet[1] if col.value == 'Water'][0].column
            medical = [col for col in sheet[1] if col.value == 'Medical'][0].column
            sanitary = [col for col in sheet[1] if col.value == 'Sanitary'][0].column
            status = [col for col in sheet[1] if col.value == 'Status'][0].column
            last_row = sheet.max_row + 1
            sheet.cell(row=last_row, column=name).value = num_people[0]
            sheet.cell(row=last_row, column=people).value = num_people[1]
            sheet.cell(row=last_row, column=lat).value = num_people[2]
            sheet.cell(row=last_row, column=lng).value = num_people[3]
            sheet.cell(row=last_row, column=food).value = num_people[4]
            sheet.cell(row=last_row, column=water).value = num_people[5]
            sheet.cell(row=last_row, column=medical).value = num_people[6]
            sheet.cell(row=last_row, column=sanitary).value = num_people[7]
            sheet.cell(row=last_row, column=status).value = num_people[8]
            workbook.save(filename)
            print("Data appended successfully!")
        except FileNotFoundError:
            print(f"File '{filename}' not found!")
        except Exception as e:
            print(f"An error occurred: {e}")




    def emergencymap(self,fullname,family,emergencyTransport,healthMedicalSupplies,criticalSituation,latitude,lognitude):
        filename = 'emergency.xlsx'
        num_people = (fullname,family,emergencyTransport,healthMedicalSupplies,criticalSituation,latitude,lognitude)
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            name = [col for col in sheet[1] if col.value == 'username'][0].column
            people = [col for col in sheet[1] if col.value == 'noofpeople'][0].column
            transport = [col for col in sheet[1] if col.value == 'transport'][0].column
            medical = [col for col in sheet[1] if col.value == 'medical'][0].column
            critical = [col for col in sheet[1] if col.value == 'critical'][0].column
            lat = [col for col in sheet[1] if col.value == 'latitude'][0].column
            lng = [col for col in sheet[1] if col.value == 'longitude'][0].column
            last_row = sheet.max_row + 1
            sheet.cell(row=last_row, column=name).value = num_people[0]
            sheet.cell(row=last_row, column=people).value = num_people[1]
            sheet.cell(row=last_row, column=transport).value = num_people[2]
            sheet.cell(row=last_row, column=medical).value = num_people[3]
            sheet.cell(row=last_row, column=critical).value = num_people[4]
            sheet.cell(row=last_row, column=lat).value = num_people[5]
            sheet.cell(row=last_row, column=lng).value = num_people[6]
            workbook.save(filename)
            print("Data appended successfully!")
        except FileNotFoundError:
            print(f"File '{filename}' not found!")
        except Exception as e:
            print(f"An error occurred: {e}")
