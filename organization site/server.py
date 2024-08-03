lat=[20.159098270646936,19.45623359601801,19.68397023588844,18.687878686034182,18.020527657852337,13.645986814875332,12.833226023521243,
    12.46876014482322,11.86735091145932,10.163560279490476,11.523087506868514,9.44906182688142,9.730714305756955,10.811724143275528]

log=[86.02294921875001,84.52880859375,83.75976562500001,84.04541015625001,83.29833984375001,75.16845703125001,75.41015625,
    76.00341796875,75.69580078125001,78.81591796875001,79.89257812500001,78.15673828125001,76.83837890625001,76.06933593750001]
name=["Aadithya", "Aarthi", "Aishwarya", "Akshaya", "Amudha",
      "Ananth", "Anjali", "Arjun", "Bala", "Bhagyalakshmi",
      "Chandra", "Dhanush", "Durga", "Ganesh", "Girija",
      "Hema", "Jagan", "Janani", "Kala", "Kamal",
      "Kanmani", "Karthik", "Kavin", "Lakshmi", "Madhan",
      "Malar", "Muthu", "Nandha", "Parthasarathy", "Priya",
      "Ramya", "Sakthi", "Sanjana", "Saranya", "Saravanan",
      "Sathya", "Selvi", "Subhash", "Sumathi", "Surya",
      "Thenmozhi", "Uma", "Vanitha", "Vijay", "Yazhini"]

import openpyxl
filename = 'emergency.xlsx'
for i in range(0,15):
    num_people = name[i]
    try:
      workbook = openpyxl.load_workbook(filename)
      sheet = workbook.active
      l_col = [col for col in sheet[1] if col.value == 'username'][0].column
      # longitude_col = [col for col in sheet[1] if col.value == 'longitude'][0].column
      last_row = sheet.max_row + 1
      sheet.cell(row=last_row, column=l_col).value = num_people
      workbook.save(filename)
      print("Data appended successfully!")
    except FileNotFoundError:
      print(f"File '{filename}' not found!")
    except Exception as e:
      print(f"An error occurred: {e}")


